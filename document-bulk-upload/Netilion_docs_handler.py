#!/usr/bin/python3

import argparse, os
import json
import requests
import openpyxl
import glob
import time

netilionUrl = "https://api.netilion.endress.com"

tokenAccess = ""
tokenRefresh = ""

documentDictionary = {}

# create datatype to pass directory path only
def dir_path(string):
    if os.path.isdir(string):
        return string
    else:
        raise NotADirectoryError(string)

def key_file(string):
    if os.path.isfile(string):
        with open(string, 'r') as file:
            data = json.load(file)
    else:
        raise FileNotFoundError
    return data

# add argument to python script
parser = argparse.ArgumentParser()
parser.add_argument("-k", dest="keys", default="test_keys.json", type=key_file, help='Credentials and Keys file path. Specific JSON format')
parser.add_argument("-d", dest="directory", type=dir_path, help='Document folder directory. Defaults to Document_Directory')
parser.add_argument("-i", dest="input", type=key_file, help='Bypass call to Netilion and supply documentDictionary as JSON')
parser.add_argument("-xl", dest="excel", type=str, help='Master file to create document Directory')

parser.add_argument("-rm", dest="remove", default=False, action='store_true', help='Delete documents from Tags in documentDictionary. Undo button.')
parser.add_argument("-mk", dest="create", default=False, action='store_true', help='No Document directory exists, so create one. Feed Excel file to create from Master File')
parser.add_argument("-p", dest="post", default=False, action='store_true', help='Upload and assign documents to tags')
parser.add_argument("-l", dest="list", default=False, action='store_true', help='list all uploaded and attached documents within documentDictionary')
parser.add_argument("-a", dest="all", default=False, action='store_true', help='Do not filter tag list to only 3rd party')

args = parser.parse_args()

if args.keys is not None:
    apiKey = args.keys["API Key"]
    apiSecret = args.keys["API Secret"]
    username = args.keys["Username"]
    password = args.keys["Password"]

if args.directory is not None:
    print("Directory provided...")
    directory = args.directory
else:
    directory = "Document_Directory"

if args.input is not None:
    print("Input bypass...")
    inputTags = args.input

if args.excel is not None:
    print("Master File directory bypass...")
    xl_directory = args.excel


removeDocs = args.remove
createDir = args.create
postDocs = args.post
listDocs = args.list
allTags = args.all


def fetchCred():
    global tokenAccess
    global tokenRefresh
    print("Querying Netilion...")

    # the request library will automatically URL percent encode the strings it is fed.
    authPayload = {'client_id': apiKey, 'client_secret': apiSecret, 'grant_type': 'password', 'username': username, 'password': password}
    authHeader = {'Content-Type': 'application/x-www-form-urlencoded'}
    response = requests.post(netilionUrl + "/oauth/token", headers=authHeader, params=authPayload)

    # response will evaluate to "true" if <399
    # so we can just raise an exception and print the response if not true.
    # some status codes in that range return something, but aren't useful.
    if response:
        authorization_dict = response.json()
        tokenAccess = authorization_dict['access_token']
        tokenRefresh = authorization_dict['refresh_token']
        print("Netilion credentials authenticated")
    else:
        raise Exception(f"Non-success status code: {response.status_code}")
    
def refreshToken():
    global tokenAccess
    global tokenRefresh

    print("Refreshing credentials...")

    authPayload = {'client_id': apiKey, 'client_secret': apiSecret, 'grant_type': 'refresh_token', 'refresh_token': tokenRefresh}
    authHeader = {'Content-Type': 'application/x-www-form-urlencoded'}
    response = requests.post(netilionUrl + "/oauth/token", headers=authHeader, params=authPayload)

    if response:
        authorization_dict = response.json()
        tokenAccess = authorization_dict['access_token']
        tokenRefresh = authorization_dict['refresh_token']
        print("Success!")
    else:
        raise Exception(f"Non-success status code: {response.status_code}")

# Query Netilion for all existing Instrumentation
# Create dictionary of Instrumentation ID and Tag
# Handle paginations
# Reject Manufacturer 63, E+H
# Assume any tag w/o an asset needs documents to upload
def pull3rdPartyTags():
    global documentDictionary

    print("Gathering tags from Netilion")

    try:
        inputTags
    except:
        ()
    else:
        documentDictionary = inputTags
        return

    response = requests.get(netilionUrl + "/v1/instrumentations?page=1&per_page=300&include=assets.product", headers={'Authorization': "Bearer " + tokenAccess})

    if response:
        response_dict = response.json()

        count = 0
        for i in response_dict["instrumentations"]:
            if response_dict["instrumentations"][count]["assets"]["total_count"] == 0:
                    documentDictionary[i["tag"]] = [i["id"], {}]
            if response_dict["instrumentations"][count]["assets"]["items"]:
                if response_dict["instrumentations"][count]["assets"]["items"][0]["product"]["manufacturer"]["id"] != 63:
                    documentDictionary[i["tag"]] = [i["id"], {}]
            count += 1

        if response_dict["pagination"]["page_count"] > 1:
            for j in range(2, 1 + response_dict["pagination"]["page_count"]):
                response = requests.get(netilionUrl + "/v1/instrumentations?page=" + str(j) + "&per_page=300", headers={'Authorization': "Bearer " + tokenAccess})
                response_dict = response.json()

                for j in response_dict["instrumentations"]:
                    documentDictionary[j["tag"]] = [j["id"], {}]
    else:
        raise Exception(f"Non-success status code: {response.status_code}")
    
    print("Created Netilion Tag list...")

    with open("logs/tag_file.log", "a") as f:
        f.write(str(documentDictionary) + "\n")
        f.close()

def pullAllTags():
    global documentDictionary

    print("Gathering ALL tags from Netilion")

    try:
        inputTags
    except:
        ()
    else:
        documentDictionary = inputTags
        return

    response = requests.get(netilionUrl + "/v1/instrumentations?page=1&per_page=300&include=assets.product", headers={'Authorization': "Bearer " + tokenAccess})

    if response:
        response_dict = response.json()

        for i in response_dict["instrumentations"]:
            documentDictionary[i["tag"]] = [i["id"], {}]

        if response_dict["pagination"]["page_count"] > 1:
            for j in range(2, 1 + response_dict["pagination"]["page_count"]):
                response = requests.get(netilionUrl + "/v1/instrumentations?page=" + str(j) + "&per_page=300", headers={'Authorization': "Bearer " + tokenAccess})
                response_dict = response.json()

                for j in response_dict["instrumentations"]:
                    documentDictionary[j["tag"]] = [j["id"], {}]
    else:
        raise Exception(f"Non-success status code: {response.status_code}")
    
    print("Created Netilion Tag list...")

    with open("logs/tag_file.log", "a") as f:
        f.write(str(documentDictionary) + "\n")
        f.close()

# Possibly create a flag to create the tags in Netilion. NOT RECOMMENDED, KEEP WITHIN CONCIERGE.
# Do SOMETHING with a supplied Excel sheet
def xlTagList():
    ()

# Determine whether files exist to upload, and how large the upload is.
def uploadSizeCalc():

    if not os.path.exists(directory):
        print("No such directory")
        exit
    
    print("Assessing content size...")

    size = 0
    for path, dirs, files in os.walk(directory):
        for f in files:
            fp = os.path.join(path, f)
            size += os.path.getsize(fp)

    # Will not let you post more than 5 GB of documents
    if size > 500000000:
        print("Too many documents")
        return False
    
    if size == 0:
        print("No files found")
        return False
    
    return True

# Append dictionary with filepaths
# Skip any tags without associated documents, or no folder for the tag in the document directory
def processFiles():

    print("Sifting through the directory for files...")
    # Write file paths within the folder to the dictionary
    folders = [x[0] for x in os.walk(directory)]

    for i in documentDictionary:
        try:
            index = [j for j, s in enumerate(folders) if i in s]
            if not index:
                documentDictionary[i][1] = ""
            else:
                documentDictionary[i][1] = glob.glob(folders[index[0]] + "*/*" , recursive=True)
        except ValueError:
            print("Files not found for specified ID...")

# Create document object, attach file to document, post document to instrumentation
# Tags will have some arbitrary number of assets
def postDocuments():

    print("Posting Documents")

    refreshCount = 0

    for i in documentDictionary:
        
        for j in documentDictionary[i][1]:
            
            if refreshCount % 200 == 0 and refreshCount > 100:
                print("\n")
                print("Refreshing Token")
                refreshToken()
                refreshCount = 0
                time.sleep(5)

            # create document
            payload = {
                "name": os.path.splitext(os.path.basename(j))[0],
                "description": str(documentDictionary[i][0]),
                "classification":{"id": 1},
                "status":{"id": 1},
                "tenant":{"id": 1}
                }
            
            # create documents
            response = requests.post(netilionUrl + "/v1/documents", headers={'Authorization': "Bearer " + tokenAccess}, json=payload)
            
            if response:
                response_dict = response.json()
                payloadDocs = {"documents": [{"id": response_dict["id"]}]}
                
                # open files to create attachment
                files = {
                    'file': (open(j, 'rb'))
                    }
                data = {'document_id': response_dict["id"]}
                
                response_Attach = requests.post(netilionUrl + "/v1/attachments", headers={'Authorization': "Bearer " + tokenAccess}, files=files, data=data)

                if response_Attach:
                    ()
                else:
                    raise Exception(f"Non-success status code: {response_Attach.status_code}")

                # assign documents
                response_Docs = requests.post(netilionUrl + "/v1/instrumentations/" + str(documentDictionary[i][0]) + "/documents", headers={'Authorization': "Bearer " + tokenAccess}, json=payloadDocs)
                
                if response_Docs:
                    # print(".", end=" ")
                    print(str(refreshCount) + " - " + str(payload))
                    refreshCount += 1
                else:
                    raise Exception(f"Non-success status code: {response_Docs.status_code}")

            else:
                raise Exception(f"Non-success status code: {response.status_code}")

    
    print("Uploaded Files to Netilion!")

# Use MIV master file to create a directory structure, if supplied as arg
# Query Netilion for Tags and create a directory structure if no arg supplied
def createDirectories():

    print("Creating document directory...")

    if not os.path.exists("Document_Directory"):
        os.makedirs("Document_Directory")

    new_directory = "Document_Directory/"

    # From File
    try:
        xl_directory
    except:
        ()
    else:
        wb = openpyxl.load_workbook(xl_directory)
        ws = wb.active

        for row in range(2, ws.max_row+1):
            dir_s1 = str(ws.cell(row, 1).value) + "/"
            dir_s2 = str(ws.cell(row, 2).value) + "/"
            dir_s3 = str(ws.cell(row, 3).value)+ "/"

            if not os.path.exists(new_directory + dir_s1):
                os.makedirs(new_directory + dir_s1)

            if not os.path.exists(new_directory + dir_s1 + dir_s2):
                os.makedirs(new_directory + dir_s1 + dir_s2)

            if not os.path.exists(new_directory + dir_s1 + dir_s2 + dir_s3):
                os.makedirs(new_directory + dir_s1 + dir_s2 + dir_s3)
        
        wb.close()
        return
    
    # From Netilion Tags
    for i in documentDictionary:
        if not os.path.exists(new_directory + i):
                os.makedirs(new_directory + i)

# utility function to list all folders within a directory that match Tag names in dict
def directoryScan():
    folders = [x[0] for x in os.walk(directory)]

    for i in documentDictionary:
        try:
            index = [j for j, s in enumerate(folders) if i in s]
            print(folders[index[0]])
        except ValueError:
            print("Not Found")

# Bulk Undo button
def deleteDocuments():

    print("Deleting all documents assigned to Tags in Tag list...")

    f = open("logs/deletion_documents.log", "a")
       
    for i in documentDictionary:
        documentList = []
        response = requests.get(netilionUrl + "/v1/instrumentations/" + str(documentDictionary[i][0]) + "/documents", headers={'Authorization': "Bearer " + tokenAccess})

        if response:
            response_dict = response.json()
            for j in response_dict["documents"]:
                documentList.append({"id": j["id"]})
        else:
            raise Exception(f"Non-success status code: {response.status_code}")

        # create the dictionary of documents
        
        data = {"documents": documentList}
        f.write(str([documentDictionary[i][0], data]) + "\n")

        if bool(documentList) is True:
            response_Deletion = requests.delete(netilionUrl + "/v1/instrumentations/" + str(documentDictionary[i][0]) + "/documents", headers={'Authorization': "Bearer " + tokenAccess, 'Content-Type': "application/json"}, data=json.dumps(data))
            
            if response_Deletion:
                ()
            else:
                raise Exception(f"Non-success status code: {response_Deletion.status_code}")
    
    f.close()
    print("Deletion Complete")

# Find any documents that were erroneously uploaded without assignment
def deleteOrphanedDocs():

    print("Finding and removing unassigned documents...")

    fullDoclist = []

    response = requests.get(netilionUrl + "/v1/documents?per_page=300", headers={'Authorization': "Bearer " + tokenAccess})

    if response:
        response_dict = response.json()
        for j in response_dict["documents"]:
            if not "document_version" in j:
                if "tenant" in j and j["tenant"]["id"] == 1:
                    fullDoclist.append(j["id"])

        if response_dict["pagination"]["page_count"] > 1:
            for j in range(2, 1 + response_dict["pagination"]["page_count"]):
                response = requests.get(netilionUrl + "/v1/documents?page=" + str(j) + "&per_page=300", headers={'Authorization': "Bearer " + tokenAccess})
                response_dict = response.json()

            for j in response_dict["documents"]:
                if not "document_version" in j:
                    if "tenant" in j and j["tenant"]["id"] == 1:
                        fullDoclist.append(j["id"])
    else:
        raise Exception(f"Non-success status code: {response.status_code}")
    
    for k in fullDoclist:
        response_delete_orphans = requests.delete(netilionUrl + "/v1/documents/" + str(k), headers={'Authorization': "Bearer " + tokenAccess, 'accept': '*/*'})
        if response_delete_orphans:
            print(str(k) + " Deleted")
        else:
            raise Exception(f"Non-success status code: {response_delete_orphans.status_code}")

    print("Deletion done.")

# Collect list of documents attached to Tags, currently
def listDocuments():
    
    print("Finding and listing documents attached to Tags in Tag List")

    f = open("logs/deletion_documents.log", "a")
    
    for i in documentDictionary:
        documentList = []
        response = requests.get(netilionUrl + "/v1/instrumentations/" + str(documentDictionary[i][0]) + "/documents", headers={'Authorization': "Bearer " + tokenAccess})

        if response:
            response_dict = response.json()
            for j in response_dict["documents"]:
                documentList.append({"id": j["id"]})
        else:
            raise Exception(f"Non-success status code: {response.status_code}")

        # create the dictionary of documents
        
        data = {"documents": documentList}
        f.write(str([documentDictionary[i][0], data]) + "\n")

    f.close()
    print("List of assigned documents written to file.")

if __name__ == "__main__":

    fetchCred()

    if not allTags:
        pull3rdPartyTags()
    else:
        pullAllTags()

    if not (removeDocs or createDir or listDocs):
        if uploadSizeCalc():
            processFiles()
            if postDocs:
                postDocuments()

    if createDir and not (removeDocs or listDocs):
        createDirectories()

    if removeDocs and not listDocs:
        deleteDocuments()

    if listDocs:
        listDocuments()

    with open("logs/runtime.log", "a") as f:
        f.write(str(documentDictionary) + "\n")
        f.close()